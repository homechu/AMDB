from collections import OrderedDict
from typing import List

from import_export.resources import ModelResource
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tablib import Dataset


class BaseResource(ModelResource):
    """基本導入導出模組

    Args:
        COLUMN_MAP (dict): 通常用於 read_only 欄位，顯示導出欄位名稱.

    Meta:
        import_id_fields (list): 判斷導入時數據，更新的唯一值.
        import_by_fields (list): 錄入導入者.
    """

    LOST_MSG = '第{}筆導入錯誤, {}欄位為必填項目'
    ERR_MSG = '第{}筆導入錯誤, {}'
    COLUMN_MAP = {}

    def get_verbose(self, field_name):
        cm = getattr(self, 'COLUMN_MAP', {})
        if cm and field_name in cm:
            v = cm[field_name]
        else:
            v = self.Meta.model._meta.get_field(field_name).verbose_name
        return v

    def get_export_headers(self):
        headers = []
        for field in self.get_export_fields():
            v = self.get_verbose(field.column_name)
            headers.append(v)
        return headers

    def get_export_order(self):
        export_include = getattr(self, 'export_include', None)
        order = tuple(self._meta.export_order or ())
        order += tuple(k for k in self.fields if k not in order)
        return (o for o in order if not export_include or o in export_include)

    def before_export(self, queryset, *args, **kwargs):
        self.export_include = kwargs.get('export_include')

    def before_import(self, dataset: Dataset, using_transactions, dry_run, **kwargs):
        fields = self.get_queryset().model._meta.get_fields()
        model_map = {i.verbose_name: i.name for i in fields if hasattr(i, 'verbose_name')}
        model_map.update({v: k for k, v in getattr(self, 'COLUMN_MAP', {}).items()})
        tmp = Dataset()
        tmp.headers = [model_map.get(i.strip(), i) for i in dataset.headers if i]
        for row in dataset:
            new_row = [
                value.strip() if isinstance(value, str) else value or ''
                for _, value in zip(tmp.headers, row)
            ]
            if any(n not in {'', None} for n in new_row):
                tmp.append(new_row)

        dataset.dict = tmp.dict

    def before_import_row(self, row: OrderedDict, row_number=None, **kwargs):
        for key in row.copy().keys():
            if key in self._meta.import_by_fields:
                row[key] = kwargs.get('username', 'system')
            elif key in self._meta.exclude + self._meta.import_exclude:
                row.pop(key)

    @staticmethod
    def xlformatter(ws: Worksheet, resouce=None):
        for c in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(c)].bestFit = True
            ws.column_dimensions[get_column_letter(c)].auto_size = True

        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if resouce and cell.value in [
                    resouce().get_verbose(r) for r in getattr(resouce.Meta, 'required', [])
                ]:
                    cell.font = Font(size=13, bold=True, color='00FF0000')
                else:
                    cell.font = Font(size=13, bold=True)

    class Meta:
        import_by_fields: List[str] = ['update_by', 'create_by']
        exclude = ('id', 'is_deleted', 'deleted_by_cascade', 'identifier')
        import_exclude = ('create_time', 'update_time')
